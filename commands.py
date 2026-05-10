import shlex

COMMANDS = {}


def register(name, help_text, schema=None):
    def decorator(fn):
        COMMANDS[name] = {"fn": fn, "help": help_text, "schema": schema or {}}
        return fn
    return decorator


def execute(session, cmd_name, **kwargs):
    entry = COMMANDS.get(cmd_name)
    if not entry:
        raise ValueError(f"Unknown command: {cmd_name}")
    return entry["fn"](session, **kwargs)


def parse(line):
    parts = shlex.split(line)
    if not parts:
        return None, {}
    cmd_name = parts[0].lower()
    kwargs = {}
    rest = parts[1:]
    i = 0
    while i < len(rest):
        arg = rest[i]
        if arg.startswith("--"):
            key = arg[2:]
            val = rest[i + 1] if i + 1 < len(rest) and not rest[i + 1].startswith("-") else True
            kwargs[key] = val
            i += 2 if val is not True else 1
        elif arg.startswith("-") and len(arg) == 2:
            key = arg[1]
            val = rest[i + 1] if i + 1 < len(rest) and not rest[i + 1].startswith("-") else True
            kwargs[key] = val
            i += 2 if val is not True else 1
        else:
            kwargs.setdefault("_pos", []).append(arg)
            i += 1
    return cmd_name, kwargs


# ── Commands ────────────────────────────────────────────


@register("cell", "Get cell value\n  cell <ref> [-s sheet] [--computed] [--style] [--table]")
def cmd_cell(session, **kwargs):
    pos = kwargs.get("_pos", [])
    if not pos:
        return {"error": "Usage: cell <ref> [-s sheet] [--computed] [--style] [--table]"}
    ref = pos[0]
    sheet = kwargs.get("s") or kwargs.get("sheet")
    try:
        val, ctype = session.cell(ref, sheet)
        result = {"ref": ref.upper(), "sheet": sheet or session.current_sheet, "value": val, "type": ctype}
        if not val:
            merged = session.cell_merged(ref, sheet)
            if merged:
                result["value"] = merged["value"]
                result["type"] = merged["type"]
                result["merged_range"] = merged["range"]
        if "computed" in kwargs:
            cv = session.cell_computed(ref, sheet)
            if cv:
                result["value"] = cv
            else:
                from openpyxl import load_workbook
                wb = load_workbook(session.path, data_only=True)
                try:
                    merged = session.cell_merged(ref, sheet, wb=wb)
                    if merged:
                        result["value"] = merged["value"]
                        result["merged_range"] = merged["range"]
                finally:
                    wb.close()
        if "style" in kwargs:
            result["style"] = session.cell_style(ref, sheet)
        if "table" in kwargs:
            result["table"] = session.cell_table(ref, sheet)
    except (ValueError, IndexError) as e:
        return {"error": str(e)}
    return result


@register("sheets", "List sheets\n  sheets [name]")
def cmd_sheets(session, **kwargs):
    pos = kwargs.get("_pos", [])
    all_sheets = session.sheets()
    if not pos:
        return all_sheets
    name = pos[0].lower()
    filtered = [s for s in all_sheets if name in s["name"].lower()]
    return filtered if filtered else {"error": f"No sheet matching '{pos[0]}'"}


@register("info", "Show file info")
def cmd_info(session, **kwargs):
    return session.info()


@register("sheet", "Switch current sheet\n  sheet <name>")
def cmd_sheet(session, **kwargs):
    pos = kwargs.get("_pos", [])
    if not pos:
        return {"error": "Usage: sheet <name>"}
    name = pos[0]
    try:
        session.switch_sheet(name)
    except ValueError as e:
        return {"error": str(e)}
    return {"sheet": name}


@register("trace", "Trace formula dependencies\n  trace <ref> [-s sheet]")
def cmd_trace(session, **kwargs):
    pos = kwargs.get("_pos", [])
    if not pos:
        return {"error": "Usage: trace <ref> [-s sheet]"}
    ref = pos[0]
    sheet = kwargs.get("s") or kwargs.get("sheet")
    try:
        return session.trace(ref, sheet)
    except ValueError as e:
        return {"error": str(e)}


@register("help", "Show this help")
def cmd_help(session, **kwargs):
    return [
        {"name": name, "help": cmd["help"].split("\n")[0]}
        for name, cmd in COMMANDS.items()
        if name != "exit"
    ]


@register("exit", "Exit the REPL")
def cmd_exit(session, **kwargs):
    raise SystemExit


@register("table", "Manage tables\n  table add <name> <range> [--header row|column|both|none]\n  table list\n  table remove <name>\n  table clear")
def cmd_table(session, **kwargs):
    pos = kwargs.get("_pos", [])
    if not pos:
        return {"error": "Usage: table add <name> <range>"}
    sub = pos[0].lower()

    if sub == "list":
        tables = session.list_tables()
        return tables if tables else {"message": "No tables registered"}

    if sub == "add":
        if len(pos) < 3:
            return {"error": "Usage: table add <name> <range>"}
        name, ref = pos[1], pos[2]
        header = kwargs.get("header", "row")
        if header not in ("row", "column", "both", "none"):
            return {"error": "Header must be row, column, both, or none"}
        try:
            session.register_table(name, ref, header=header)
        except ValueError as e:
            return {"error": str(e)}
        return {"message": f"Registered table \"{name}\" ({ref.upper()}) with header={header}", "name": name, "range": ref.upper(), "header": header}

    if sub == "remove":
        if len(pos) < 2:
            return {"error": "Usage: table remove <name>"}
        name = pos[1]
        if session.unregister_table(name):
            return {"message": f"Removed table \"{name}\""}
        return {"error": f"Table \"{name}\" not found"}

    if sub == "clear":
        session.clear_tables()
        return {"message": "All tables cleared"}

    return {"error": f"Unknown subcommand: {sub}"}
