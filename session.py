import json
import os
import re
from datetime import datetime, date

from openpyxl import load_workbook, utils

_TYPE_LABELS = {
    "n": "number",
    "s": "text",
    "f": "formula",
    "d": "date",
    "b": "boolean",
    "e": "error",
    "nul": "empty",
}


class Session:
    def __init__(self, path):
        if not os.path.exists(path):
            raise FileNotFoundError(f"File not found: {path}")
        self.path = path
        self.wb = load_workbook(path, data_only=False)
        self._current_sheet = self.wb.sheetnames[0]
        self._tables_path = path + ".tables.json"
        self._tables = {}
        self._load_tables()

    @property
    def current_sheet(self):
        return self._current_sheet

    def close(self):
        self.wb.close()

    def info(self):
        return {
            "filename": os.path.basename(self.path),
            "size": os.path.getsize(self.path),
            "sheets": len(self.wb.sheetnames),
        }

    def sheets(self):
        result = []
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            result.append({
                "name": name,
                "hidden": ws.sheet_state == "hidden",
                "rows": ws.max_row or 0,
                "cols": len(first),
            })
        return result

    def switch_sheet(self, name):
        if name not in self.wb.sheetnames:
            raise ValueError(
                f"Sheet '{name}' not found. Available: {', '.join(self.wb.sheetnames)}"
            )
        self._current_sheet = name

    def cell(self, ref, sheet=None):
        cell = self._resolve_cell(ref, sheet)
        return self._fmt(cell.value), _TYPE_LABELS.get(cell.data_type, "unknown")

    def cell_computed(self, ref, sheet=None):
        sheet = sheet or self._current_sheet
        wb = load_workbook(self.path, data_only=True)
        try:
            ws = wb[sheet]
            cell = self._resolve_cell(ref, sheet, wb)
            return self._fmt(cell.value)
        finally:
            wb.close()

    def cell_merged(self, ref, sheet=None, wb=None):
        sheet = sheet or self._current_sheet
        wb = wb or self.wb
        ws = wb[sheet]
        cell = self._resolve_cell(ref, sheet, wb)
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                top = ws.cell(mr.min_row, mr.min_col)
                return {
                    "range": str(mr),
                    "value": self._fmt(top.value),
                    "type": _TYPE_LABELS.get(top.data_type, "unknown"),
                }
        return None

    def trace(self, ref, sheet=None):
        sheet = sheet or self._current_sheet
        val, ctype = self.cell(ref, sheet)
        if ctype != "formula":
            merged = self.cell_merged(ref, sheet)
            if merged and merged["type"] == "formula":
                mrange = merged["range"]
                min_col, min_row, _, _ = utils.range_boundaries(mrange)
                ref = f"{utils.get_column_letter(min_col)}{min_row}"
                cell = self._resolve_cell(ref, sheet)
                val, ctype = self.cell(ref, sheet)
            else:
                raise ValueError(f"'{ref}' is not a formula")
        else:
            cell = self._resolve_cell(ref, sheet)
        formula = cell.value or ""
        pattern = r"(?:(?:'[^']+'|[A-Za-z_]\w*)!)?\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?"
        matches = re.findall(pattern, formula)
        seen = set()
        deps = []
        for m in matches:
            if "!" in m:
                s_part, r_part = m.split("!", 1)
                s_name = s_part.strip("'")
                r_clean = r_part.replace("$", "")
            else:
                s_name = sheet
                r_clean = m.replace("$", "")
            key = f"{s_name}!{r_clean}"
            if key in seen:
                continue
            seen.add(key)
            deps.append({"ref": r_clean, "sheet": s_name})
        results = []
        for dep in deps:
            if ":" in dep["ref"]:
                results.append({"ref": dep["ref"], "sheet": dep["sheet"], "value": None, "type": "range", "computed": None})
            else:
                try:
                    v, t = self.cell(dep["ref"], dep["sheet"])
                except ValueError:
                    v = "(error)"
                    t = "error"
                results.append({"ref": dep["ref"], "sheet": dep["sheet"], "value": v, "type": t, "computed": None})
        computed = self.cell_computed(ref, sheet)
        return {
            "ref": ref.upper(),
            "sheet": sheet,
            "formula": str(formula),
            "computed": computed,
            "dependencies": results,
        }

    def find(self, query=None, sheet=None, type_filter=None, formula_pattern=None, merged_only=False):
        sheet = sheet or self._current_sheet
        ws = self.wb[sheet]
        q = query.lower() if query else None
        fp = formula_pattern.lower() if formula_pattern else None
        _type_map = {"number": "n", "text": "s", "formula": "f"}
        tf = _type_map.get(type_filter) if type_filter else None
        merged_ranges = list(ws.merged_cells.ranges) if merged_only else None
        results = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if merged_only:
                    in_merged = any(cell.coordinate in mr for mr in merged_ranges)
                    if not in_merged:
                        continue
                if tf and cell.data_type != tf:
                    continue
                if type_filter == "empty" and cell.value is not None:
                    continue
                if fp:
                    if cell.data_type != "f":
                        continue
                    raw = str(cell.value or "")
                    if fp not in raw.lower():
                        continue
                if q:
                    val_str = str(cell.value or "").lower()
                    ref_str = cell.coordinate.lower()
                    if q not in val_str and q not in ref_str:
                        continue
                merged_range = None
                if merged_ranges:
                    for mr in merged_ranges:
                        if cell.coordinate in mr:
                            merged_range = str(mr)
                            break
                val = self._fmt(cell.value)
                if not val and cell.data_type != "nul" and type_filter != "empty":
                    merged = self.cell_merged(cell.coordinate, sheet)
                    if merged:
                        val = merged["value"]
                        merged_range = merged_range or merged["range"]
                results.append({
                    "ref": cell.coordinate,
                    "sheet": sheet,
                    "value": val,
                    "type": _TYPE_LABELS.get(cell.data_type, "unknown"),
                    "merged_range": merged_range,
                })
        return results

    def cell_style(self, ref, sheet=None):
        cell = self._resolve_cell(ref, sheet)
        f = cell.font
        fi = cell.fill
        b = cell.border
        a = cell.alignment
        return {
            "font": {
                "bold": f.bold,
                "italic": f.italic,
                "size": f.size,
                "name": f.name,
                "color": str(f.color.rgb) if f.color and f.color.rgb else None,
            },
            "fill": {
                "fg": str(fi.fgColor.rgb) if fi.fgColor and fi.fgColor.rgb else None,
                "bg": str(fi.bgColor.rgb) if fi.bgColor and fi.bgColor.rgb else None,
            },
            "border": {
                "top": b.top.style if b.top else None,
                "bottom": b.bottom.style if b.bottom else None,
                "left": b.left.style if b.left else None,
                "right": b.right.style if b.right else None,
            },
            "align": {
                "h": a.horizontal if a else None,
                "v": a.vertical if a else None,
                "wrap": a.wrap_text if a else None,
            },
            "number_format": cell.number_format,
        }

    def cell_table(self, ref, sheet=None):
        sheet = sheet or self._current_sheet
        ws = self.wb[sheet]
        col, row = self._parse_ref(ref)

        def _resolve(val, r, c):
            if val != "":
                return val
            m = self.cell_merged(f"{utils.get_column_letter(c)}{r}", sheet, wb=self.wb)
            return m["value"] if m else ""

        for tname, tref in ws.tables.items():
            bounds = utils.range_boundaries(tref)
            if bounds and bounds[0] <= col <= bounds[2] and bounds[1] <= row <= bounds[3]:
                header_cell = ws.cell(row=bounds[1], column=col)
                return {
                    "name": tname,
                    "range": tref,
                    "column": _resolve(self._fmt(header_cell.value), bounds[1], col),
                }
        for tname, meta in self._tables.get(sheet, {}).items():
            if isinstance(meta, str):
                continue
            bounds = utils.range_boundaries(meta["range"])
            if bounds and bounds[0] <= col <= bounds[2] and bounds[1] <= row <= bounds[3]:
                result = {"name": tname, "range": meta["range"]}
                h = meta.get("header", "row")
                if h in ("row", "both"):
                    result["column"] = _resolve(self._fmt(ws.cell(row=bounds[1], column=col).value), bounds[1], col)
                if h in ("column", "both"):
                    result["row"] = _resolve(self._fmt(ws.cell(row=row, column=bounds[0]).value), row, bounds[0])
                return result
        return None

    def register_table(self, name, ref, sheet=None, header="row"):
        sheet = sheet or self._current_sheet
        try:
            utils.range_boundaries(ref)
        except (ValueError, AttributeError):
            raise ValueError(f"Invalid range: {ref}")
        self._tables.setdefault(sheet, {})[name] = {"range": ref.upper(), "header": header}
        self._save_tables()

    def unregister_table(self, name, sheet=None):
        sheet = sheet or self._current_sheet
        result = self._tables.get(sheet, {}).pop(name, None) is not None
        if result:
            self._save_tables()
        return result

    def list_tables(self, sheet=None):
        sheet = sheet or self._current_sheet
        result = []
        for name, meta in self._tables.get(sheet, {}).items():
            if isinstance(meta, str):
                continue
            result.append({"name": name, "range": meta["range"], "header": meta.get("header", "row"), "sheet": sheet})
        return result

    def clear_tables(self, sheet=None):
        sheet = sheet or self._current_sheet
        self._tables.pop(sheet, None)
        self._save_tables()

    def _load_tables(self):
        try:
            with open(self._tables_path) as f:
                raw = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self._tables = {}
            return
        self._tables = {}
        for sheet, tables in raw.items():
            if not isinstance(tables, dict):
                continue
            self._tables[sheet] = {}
            for name, meta in tables.items():
                if isinstance(meta, dict) and "range" in meta:
                    self._tables[sheet][name] = meta

    def _save_tables(self):
        with open(self._tables_path, "w") as f:
            json.dump(self._tables, f, indent=2)

    def _resolve_cell(self, ref, sheet, wb=None):
        wb = wb or self.wb
        sheet = sheet or self._current_sheet
        if sheet not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found.")
        ws = wb[sheet]
        col, row = self._parse_ref(ref)
        return ws.cell(row=row, column=col)

    @staticmethod
    def _parse_ref(ref):
        ref = ref.upper()
        col_str = "".join(c for c in ref if c.isalpha())
        row_str = "".join(c for c in ref if c.isdigit())
        if not col_str or not row_str:
            raise ValueError(f"Invalid cell reference: {ref}")
        return utils.column_index_from_string(col_str), int(row_str)

    @staticmethod
    def _fmt(v):
        if v is None:
            return ""
        if isinstance(v, (datetime, date)):
            return v.isoformat()
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v)
